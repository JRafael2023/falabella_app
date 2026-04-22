from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
ws.title = "Hallazgos"

headers = [
    "titulo", "gerencia", "ecosistema", "gerente", "auditor",
    "nivel_riesgo", "estado_publicacion", "tipo_impacto",
    "soporte_ecosistema", "tipo_riesgo", "tipologia_riesgo", "alcance_observacion"
]

data = [
    ["Fraude en transacciones digitales", "Gerencia de Riesgo", "Ecosistema Digital",
     "Juan Pérez", "María López", "Alto", "Publicado", "Financiero",
     "Soporte Crítico", "Riesgo Operacional", "Fraude Externo", "Nacional"],
    ["Incumplimiento de políticas de crédito", "Gerencia Comercial", "Ecosistema Retail",
     "Carlos Gómez", "Ana Torres", "Medio", "Borrador", "Reputacional",
     "Soporte Estándar", "Riesgo Crediticio", "Mora Temprana", "Regional"],
    ["Pérdida de inventario en bodega", "Gerencia de Operaciones", "Ecosistema Supply Chain",
     "Pedro Ruiz", "Sofía Vargas", "Bajo", "En Revisión", "Operacional",
     "Soporte Básico", "Riesgo Logístico", "Merma Interna", "Local"],
]

dark_blue = PatternFill("solid", start_color="FF1F4E79", end_color="FF1F4E79")
med_blue  = PatternFill("solid", start_color="FF2E75B6", end_color="FF2E75B6")
yellow    = PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
no_fill   = PatternFill(fill_type=None)

white_bold  = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
arial_data  = Font(name="Arial", size=11)

thin = Side(style="thin", color="FF000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write headers row 1
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = white_bold
    cell.fill = dark_blue if col_idx <= 5 else med_blue
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Write data rows 2-4
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = arial_data
        cell.fill = yellow if col_idx <= 5 else no_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = border

# Column widths
for col_letter in "ABCDEFGHIJKL":
    ws.column_dimensions[col_letter].width = 25

# Row 1 height
ws.row_dimensions[1].height = 25

# Comment on A1
comment_text = "OBLIGATORIAS: A-E (titulo, gerencia, ecosistema, gerente, auditor)\nOPCIONALES: F-L"
comment = Comment(comment_text, "Template")
ws["A1"].comment = comment

wb.save(r"C:\Users\USER\Downloads\falabella\template_importar_hallazgos.xlsx")
print("Saved.")
